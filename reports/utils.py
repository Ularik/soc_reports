from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from django.contrib.auth import get_user_model
from user.models import CustomUser
from io import BytesIO
from openpyxl.utils import get_column_letter
from copy import copy

User = get_user_model()


def create_stat_report(data, start, end):
    workbook = load_workbook('reports/stat_rep_template/pattern.xlsx')
    ws = workbook['Лист1']  # Указываем имя листа

    start_day, start_month = start.day, start.month
    end_day, end_month = end.day, end.month
    year = end.year

    ws['A2'] = (f'СТАТИСТИЧЕСКАЯ ФОРМА ОТЧЕТНОСТИ по мониторингу SOC "1" отдела КЦОКБ за период с '
                f'{start_day}.{start_month}.{year} по {end_day}.{end_month}.{year}')

    users = CustomUser.objects.all()

    start_row = 6  # внутри excel начинаем заполнять поля с 6 строки

    for i, usr in enumerate(users, 1):
        row = start_row + i    # номер строки
        if i > 4:
            ws.insert_rows(row)

            # Копируем стиль со строки выше
            for col in range(1, ws.max_column + 1):
                col_letter = get_column_letter(col)
                cell_above = ws[f'{col_letter}{row - 1}']
                cell_new = ws[f'{col_letter}{row}']
                cell_new._style = copy(cell_above._style)

        ws['A' + str(row)] = i  # заполняем числами нумерацию в первом столбце
        ws['B' + str(row)] = usr.username_ru if usr.username_ru else usr.username  # заполняем имя пользователя во втором столбце

        organizations_dct = data.get(usr.username)
        if organizations_dct is None:
            for symbol_column in ('D', 'E', 'F', 'G', 'H', 'I', 'N', 'O'):
                ws[f"{symbol_column}{row}"] = 0
        else:
            for org in organizations_dct:
                risk_assessment_dict = organizations_dct[org]

                if org == 'FIN':
                    ws[f'D{row}'] = risk_assessment_dict.get('Критическая', 0) + risk_assessment_dict.get('Высокая', 0)
                    ws[f'E{row}'] = risk_assessment_dict.get('Средняя', 0) + risk_assessment_dict.get('Низкая', 0)
                elif org == 'GTS':
                    ws[f'F{row}'] = risk_assessment_dict.get('Критическая', 0) + risk_assessment_dict.get('Высокая', 0)
                    ws[f'G{row}'] = risk_assessment_dict.get('Средняя', 0) + risk_assessment_dict.get('Низкая', 0)
                elif org == 'GNS':
                    ws[f'H{row}'] = risk_assessment_dict.get('Критическая', 0) + risk_assessment_dict.get('Высокая', 0)
                    ws[f'I{row}'] = risk_assessment_dict.get('Средняя', 0) + risk_assessment_dict.get('Низкая', 0)
                elif org =='ADM':
                    ws[f'N{row}'] = risk_assessment_dict.get('Критическая', 0) + risk_assessment_dict.get('Высокая', 0)
                    ws[f'O{row}'] = risk_assessment_dict.get('Средняя', 0) + risk_assessment_dict.get('Низкая', 0)

    row_for_results = start_row + 1 + len(users)   # строка для подсчета итогов всех строк
    # Суммирование по столбцам
    for col in range(3, 21):    # 3, 21 столбцы с количеством отчетов
        col_letter = get_column_letter(col)
        formula = f"=SUM({col_letter}{start_row + 1}:{col_letter}{start_row + len(users)})"
        ws.cell(row=row_for_results, column=col).value = formula

    # Сохранение файла
    # Сохранение файла в память
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output
